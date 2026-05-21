"""Sığorta Hesabatı Generatoru — Flask web app.

Upload a hospital services .xlsx, group rows by column N (Müəssisə), and
download per-insurance reports containing only the chosen output columns.
"""
from __future__ import annotations

import io
import os
import re
import uuid
import zipfile
from datetime import datetime, date
from threading import Lock

import pandas as pd
from rapidfuzz import fuzz, process
from flask import (
    Flask,
    Response,
    abort,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 250 * 1024 * 1024  # 250 MB cap

# Set APP_PASSWORD in the environment to require a password to access the app.
# If unset, the app is open (useful for local development only).
APP_PASSWORD = os.environ.get("APP_PASSWORD", "").strip() or None


@app.before_request
def _require_password():
    if APP_PASSWORD is None:
        return None
    auth = request.authorization
    if auth is not None and auth.password == APP_PASSWORD:
        return None
    return Response(
        "Daxil olmaq üçün parol tələb olunur.\n",
        401,
        {"WWW-Authenticate": 'Basic realm="Sigorta Hesabati"'},
    )


# ----------------------- Column configuration -----------------------
#
# Columns are located in the source file by HEADER NAME, not by fixed position.
# This makes the app robust to files where columns have been inserted, removed,
# or reordered. Each header is normalized (strip + collapse whitespace +
# casefold) before lookup, so minor formatting differences don't break things.
#
# We read these source columns: Soyadı, Adı, Baba Adı are used as inputs for
# the computed Tam adi (and Adı is used as a sort key); they're kept in the
# output but marked HIDDEN. Tam adi is computed by us — we never trust a
# pre-existing "Tam adi" column in the source.

# normalized-header  ->  internal role name
TARGET_HEADERS: dict[str, str] = {
    "xidmət tarixi":    "date_service",
    "protokol no":      "protocol",
    "soyadı":           "soyadi",
    "adı":              "adi",
    "baba adı":         "baba",
    "doğum tarixi":     "dob",
    "müəssisə":         "group",
    "xidmət adı":       "service",
    "ədəd":             "qty",
    "kdvli hasta tutar": "amount",
}
REQUIRED_ROLES = list(TARGET_HEADERS.values())
ROLE_TO_LABEL = {role: label for label, role in TARGET_HEADERS.items()}

TAM_ADI_LABEL = "Tam adi"  # computed column

UNKNOWN_LABEL = "Bilinmir"  # used when Müəssisə is empty/NaN

# Price-list matching configuration.
# File 2 (price list) shape:
#   - Header row at row 4 (1-indexed) → pandas header=3
#   - Data starts at row 5
#   - Column D (index 3) = service name to match against
#   - Column E (index 4) = price to copy into the output
PRICE_LIST_HEADER_ROW = 3            # 0-indexed
PRICE_LIST_NAME_COL_IDX = 3          # column D
PRICE_LIST_PRICE_COL_IDX = 4         # column E
MATCH_THRESHOLD = 70.0               # rapidfuzz score >= 70 counts as a match
MATCHED_NAME_LABEL = "Xidmətin adı (sığorta)"
MATCHED_PRICE_LABEL = "Qiymət"


# ----------------------- Azerbaijani alphabet sort -----------------------

AZ_LETTERS = [
    ("A", "a"), ("B", "b"), ("C", "c"), ("Ç", "ç"), ("D", "d"), ("E", "e"),
    ("Ə", "ə"), ("F", "f"), ("G", "g"), ("Ğ", "ğ"), ("H", "h"), ("X", "x"),
    ("I", "ı"), ("İ", "i"), ("J", "j"), ("K", "k"), ("Q", "q"), ("L", "l"),
    ("M", "m"), ("N", "n"), ("O", "o"), ("Ö", "ö"), ("P", "p"), ("R", "r"),
    ("S", "s"), ("Ş", "ş"), ("T", "t"), ("U", "u"), ("Ü", "ü"), ("V", "v"),
    ("Y", "y"), ("Z", "z"),
]

_AZ_RANK: dict[str, int] = {}
for _i, (_u, _l) in enumerate(AZ_LETTERS):
    _AZ_RANK[_u] = _i
    _AZ_RANK[_l] = _i


def _az_sort_key(s: str) -> str:
    """Map a string to a comparable key string that sorts by the Azerbaijani alphabet."""
    if not s:
        return ""
    parts = []
    for ch in s:
        rank = _AZ_RANK.get(ch)
        if rank is None:
            # Non-alphabet chars (spaces, digits, punctuation, foreign letters)
            # sort after all alphabet chars; preserve their relative order.
            rank = 100 + (ord(ch) & 0x3FF)
        parts.append(f"{rank:04x}")
    return "".join(parts)


# ----------------------- Header normalization -----------------------


def _norm_header(h) -> str:
    """Casefold + collapse whitespace so 'Xidmət  adı', ' xidmət adı ' etc. all
    normalize to 'xidmət adı'."""
    if h is None:
        return ""
    return re.sub(r"\s+", " ", str(h)).strip().casefold()


# ----------------------- Storage -----------------------

# In-memory store keyed by upload id:
#   {upload_id: {"filename": str, "groups": {name: DataFrame},
#                "col_types": {col_name: 'date'|'money'|'qty'|'text'},
#                "created": datetime}}
STORE: dict[str, dict] = {}
STORE_LOCK = Lock()


# ----------------------- Date parsing -----------------------

_DATE_PATTERNS = [
    "%Y-%m-%d %H:%M:%S.%f",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d",
    "%d/%m/%Y %H:%M:%S",
    "%d/%m/%Y",
    "%d.%m.%Y",
    "%d-%m-%Y",
]


def _try_parse_date(s: str):
    s = s.strip()
    if not s:
        return None
    for fmt in _DATE_PATTERNS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _normalize_to_datetime(v):
    """Coerce a value to a pandas Timestamp (or NaT) for sorting/formatting."""
    if v is None:
        return pd.NaT
    if isinstance(v, float) and pd.isna(v):
        return pd.NaT
    if isinstance(v, pd.Timestamp):
        return v
    if isinstance(v, datetime):
        return pd.Timestamp(v)
    if isinstance(v, date):
        return pd.Timestamp(v)
    if isinstance(v, str):
        parsed = _try_parse_date(v)
        if parsed is not None:
            return pd.Timestamp(parsed)
        return pd.NaT
    return pd.NaT


# ----------------------- Price-list parsing & fuzzy matching -----------------------


def _normalize_service_name(s) -> str:
    """Strip whitespace/newlines, collapse internal whitespace, casefold."""
    if s is None:
        return ""
    if isinstance(s, float) and pd.isna(s):
        return ""
    s = str(s).strip()
    if not s:
        return ""
    s = re.sub(r"\s+", " ", s)
    return s.casefold()


def _parse_price_list(file_storage):
    """Parse the price list. Returns (normalized_names, raw_names, prices) lists.

    Layout: header at row 4 (1-indexed), data from row 5. We read columns D + E
    only. Names are kept verbatim (for display) plus a normalized copy (for
    matching). Empty rows are dropped.
    """
    df = pd.read_excel(
        file_storage,
        header=PRICE_LIST_HEADER_ROW,
        usecols=[PRICE_LIST_NAME_COL_IDX, PRICE_LIST_PRICE_COL_IDX],
        engine="openpyxl",
    )
    name_col = df.columns[0]
    price_col = df.columns[1]

    names_norm: list[str] = []
    names_raw: list[str] = []
    prices: list = []
    for raw_name, raw_price in zip(df[name_col].tolist(), df[price_col].tolist()):
        norm = _normalize_service_name(raw_name)
        if not norm:
            continue
        names_norm.append(norm)
        names_raw.append(str(raw_name).strip())
        prices.append(raw_price)
    return names_norm, names_raw, prices


def _match_services(
    service_values: list,
    price_list: tuple[list[str], list[str], list],
    threshold: float = MATCH_THRESHOLD,
) -> dict[str, tuple[str, object]]:
    """For each unique service name, find the best fuzzy match above threshold.

    Returns {original_service_name: (matched_raw_name, matched_price)}. Names
    with no match (or empty) are absent from the dict.
    """
    names_norm, names_raw, prices = price_list
    if not names_norm:
        return {}

    seen = set()
    unique: list = []
    for v in service_values:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        if v in seen:
            continue
        seen.add(v)
        unique.append(v)

    match_map: dict[str, tuple[str, object]] = {}
    for src in unique:
        src_norm = _normalize_service_name(src)
        if not src_norm:
            continue
        result = process.extractOne(
            src_norm,
            names_norm,
            scorer=fuzz.ratio,
            score_cutoff=threshold,
        )
        if result is None:
            continue
        _, _, idx = result
        match_map[src] = (names_raw[idx], prices[idx])
    return match_map


# ----------------------- Match-existing-file flow (box 2) -----------------------


# Header names we recognize inside an existing (already-split) insurance file.
# Matching is case-insensitive and whitespace-tolerant.
_SERVICE_HEADER = "xidmət adı"
_HIDDEN_HEADER_NAMES = {"soyadı", "adı", "baba adı"}
_DATE_HEADER_NAMES = {"xidmət tarixi", "doğum tarixi"}
_MONEY_HEADER_NAMES = {"kdvli hasta tutar", "qiymət"}
_QTY_HEADER_NAMES = {"ədəd"}


def _match_existing_file(insurance_storage, pricelist_storage) -> tuple[bytes, str]:
    """Take an already-split insurance .xlsx + a price list, return new .xlsx
    bytes (and a suggested download name) with two new columns appended:
    "Xidmətin adı (sığorta)" and "Qiymət".
    """
    df = pd.read_excel(insurance_storage, engine="openpyxl")

    # Find the Xidmət adı column by header name.
    service_col = None
    for col in df.columns:
        if _norm_header(col) == _SERVICE_HEADER:
            service_col = col
            break
    if service_col is None:
        raise ValueError(
            "'Xidmət adı' sütunu bu faylda tapılmadı. Zəhmət olmasa "
            "1-ci addımın çıxış faylını yükləyin."
        )

    price_list = _parse_price_list(pricelist_storage)
    match_map = _match_services(df[service_col].tolist(), price_list)

    df[MATCHED_NAME_LABEL] = df[service_col].map(
        lambda v: match_map.get(v, (None, None))[0]
    )
    df[MATCHED_PRICE_LABEL] = df[service_col].map(
        lambda v: match_map.get(v, (None, None))[1]
    )

    # Classify each column for formatting + figure out which to hide.
    col_types: dict[str, str] = {}
    hidden_cols: set[str] = set()
    for col in df.columns:
        n = _norm_header(col)
        if n in _DATE_HEADER_NAMES:
            col_types[col] = "date"
        elif n in _MONEY_HEADER_NAMES:
            col_types[col] = "money"
        elif n in _QTY_HEADER_NAMES:
            col_types[col] = "qty"
        else:
            col_types[col] = "text"
        if n in _HIDDEN_HEADER_NAMES:
            hidden_cols.add(col)

    data = _df_to_xlsx_bytes(df, col_types, hidden_cols, sheet_name="Hesabat")
    return data


# ----------------------- Core: read, sort, group -----------------------


def _locate_columns(df) -> dict[str, str]:
    """Map each TARGET_HEADERS role to the actual column name found in df.

    Raises ValueError listing the missing headers (in human-readable form) if
    any required column can't be found.
    """
    found: dict[str, str] = {}
    for col in df.columns:
        n = _norm_header(col)
        if n in TARGET_HEADERS:
            role = TARGET_HEADERS[n]
            # First match wins (in case of duplicate headers).
            found.setdefault(role, col)
    missing = [role for role in REQUIRED_ROLES if role not in found]
    if missing:
        missing_labels = [ROLE_TO_LABEL[r] for r in missing]
        raise ValueError(
            "Tələb olunan sütun(lar) tapılmadı: "
            + ", ".join(repr(lbl) for lbl in missing_labels)
            + ". Excel faylındakı sütun başlıqlarını yoxlayın."
        )
    return found


def _read_and_group(
    file_storage,
) -> tuple[dict[str, pd.DataFrame], dict[str, str], set[str]]:
    """Read the uploaded xlsx, sort, group by Müəssisə.

    Columns are located by HEADER NAME (not by fixed position) so files with
    inserted/removed columns still work. Output keeps the source Soyadı/Adı/
    Baba Adı (hidden in Excel) plus a computed visible "Tam adi". Price-list
    matching is a separate flow — see /match.
    """
    df = pd.read_excel(file_storage, engine="openpyxl")
    cols = _locate_columns(df)

    date_col = cols["date_service"]
    dob_col = cols["dob"]
    soyadi_col = cols["soyadi"]
    adi_col = cols["adi"]
    baba_col = cols["baba"]
    group_col = cols["group"]

    # Build computed Tam adi = Soyadı + Adı + Baba Adı (vectorized).
    df[TAM_ADI_LABEL] = (
        df[soyadi_col].fillna("").astype(str).str.strip()
        + " " + df[adi_col].fillna("").astype(str).str.strip()
        + " " + df[baba_col].fillna("").astype(str).str.strip()
    ).str.replace(r"\s+", " ", regex=True).str.strip()

    # Normalize date columns so sorting / writing have real datetimes.
    df[date_col] = df[date_col].apply(_normalize_to_datetime)
    df[dob_col] = df[dob_col].apply(_normalize_to_datetime)

    # Sort: primary Xidmət tarixi ascending (oldest first),
    #       secondary Adı ascending using Azerbaijani alphabet.
    df["_az_sort_key"] = df[adi_col].fillna("").astype(str).map(_az_sort_key)
    df = df.sort_values(
        by=[date_col, "_az_sort_key"],
        ascending=[True, True],
        na_position="last",
        kind="mergesort",  # stable
    ).reset_index(drop=True)
    df = df.drop(columns=["_az_sort_key"])

    # Output column order: A, D, E (hidden), F (hidden), G (hidden), Tam adi,
    # Doğum Tarixi, Xidmət adı, Ədəd, Kdvli Hasta Tutar.
    output_cols = [
        date_col,
        cols["protocol"],
        soyadi_col,
        adi_col,
        baba_col,
        TAM_ADI_LABEL,
        dob_col,
        cols["service"],
        cols["qty"],
        cols["amount"],
    ]
    hidden_cols = {soyadi_col, adi_col, baba_col}

    col_types: dict[str, str] = {
        date_col: "date",
        cols["protocol"]: "text",
        soyadi_col: "text",
        adi_col: "text",
        baba_col: "text",
        TAM_ADI_LABEL: "text",
        dob_col: "date",
        cols["service"]: "text",
        cols["qty"]: "qty",
        cols["amount"]: "money",
    }

    # Group by Müəssisə.
    groups: dict[str, pd.DataFrame] = {}
    for name, sub in df.groupby(group_col, dropna=False, sort=True):
        if pd.isna(name) or str(name).strip() == "":
            key = UNKNOWN_LABEL
        else:
            key = str(name).strip()
        sub = sub[output_cols].reset_index(drop=True)
        if key in groups:
            groups[key] = pd.concat([groups[key], sub], ignore_index=True)
        else:
            groups[key] = sub
    return groups, col_types, hidden_cols


# ----------------------- Display formatting (HTML preview) -----------------------


def _fmt_cell(col_type: str, val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    if pd.isna(val):
        return ""
    if col_type == "date":
        if isinstance(val, (pd.Timestamp, datetime, date)):
            return val.strftime("%d.%m.%Y")
        s = str(val).strip()
        if " " in s and re.match(r"^\d{4}-\d{2}-\d{2}", s):
            s = s.split(" ", 1)[0]
        return s
    if col_type == "money":
        try:
            return f"{float(val):,.2f}"
        except (TypeError, ValueError):
            return str(val)
    if col_type == "qty":
        try:
            f = float(val)
            return str(int(f)) if f.is_integer() else f"{f:g}"
        except (TypeError, ValueError):
            return str(val)
    return str(val)


# ----------------------- Excel writing -----------------------


def _df_to_xlsx_bytes(
    df: pd.DataFrame,
    col_types: dict[str, str],
    hidden_cols: set[str] | None = None,
    sheet_name: str = "Sheet1",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name(sheet_name, set())
    _write_df_to_sheet(ws, df, col_types, hidden_cols or set())
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Excel table style options:
#   TableStyleLight* (subtle), TableStyleMedium* (default-ish), TableStyleDark*
# "TableStyleMedium2" gives a clean yellow/gold accent + banded rows + filter buttons.
_TABLE_STYLE = "TableStyleMedium2"

# Border for non-table fallback (currently every output uses a Table, so this
# is just defensive in case `add_table` ever fails).
_THIN = Side(border_style="thin", color="BFBFBF")
_CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _write_df_to_sheet(
    ws,
    df: pd.DataFrame,
    col_types: dict[str, str],
    hidden_cols: set[str] | None = None,
) -> None:
    hidden_cols = hidden_cols or set()
    headers = list(df.columns)
    ws.append(headers)

    # Write data rows.
    for row in df.itertuples(index=False, name=None):
        out = []
        for col_name, val in zip(headers, row):
            ctype = col_types.get(col_name, "text")
            if val is None or (isinstance(val, float) and pd.isna(val)):
                out.append(None)
            elif ctype == "date":
                if isinstance(val, pd.Timestamp):
                    out.append(val.to_pydatetime())
                elif isinstance(val, (datetime, date)):
                    out.append(val)
                else:
                    parsed = _try_parse_date(str(val))
                    out.append(parsed if parsed is not None else val)
            else:
                out.append(val)
        ws.append(out)

    n_rows = ws.max_row
    n_cols = len(headers)
    last_col_letter = get_column_letter(n_cols)

    # Number formats per column type.
    money_fmt = "#,##0.00"
    date_fmt = "dd.mm.yyyy"
    for col_idx, name in enumerate(headers, start=1):
        ctype = col_types.get(name, "text")
        if ctype == "date":
            for r in range(2, n_rows + 1):
                ws.cell(row=r, column=col_idx).number_format = date_fmt
        elif ctype == "money":
            for r in range(2, n_rows + 1):
                ws.cell(row=r, column=col_idx).number_format = money_fmt
        elif ctype == "qty":
            for r in range(2, n_rows + 1):
                ws.cell(row=r, column=col_idx).number_format = "0"

    # Alignment for header row (the Table style will handle colors/borders).
    for col_idx in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col_idx)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-fit-ish column widths (sample up to first 200 rows; cap to keep sane).
    for col_idx, name in enumerate(headers, start=1):
        max_len = len(str(name))
        for r in range(2, min(n_rows, 201) + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                continue
            if isinstance(v, (datetime, date)):
                s = v.strftime("%d.%m.%Y")
            else:
                s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 45)

    # Hide source-name columns (E/F/G equivalents). The data is still there;
    # users can right-click and "Unhide" if they need to see the components.
    for col_idx, name in enumerate(headers, start=1):
        if name in hidden_cols:
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    # Apply Excel "Format as Table" — gives borders, banded rows, filter dropdowns.
    if n_rows >= 2:  # need at least 1 data row for a valid table range
        table_ref = f"A1:{last_col_letter}{n_rows}"
        try:
            tab = Table(displayName="Hesabat", ref=table_ref)
            tab.tableStyleInfo = TableStyleInfo(
                name=_TABLE_STYLE,
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(tab)
        except Exception:
            # Fallback: at least draw plain borders so the output still has lines.
            for r in range(1, n_rows + 1):
                for c in range(1, n_cols + 1):
                    ws.cell(row=r, column=c).border = _CELL_BORDER

    ws.freeze_panes = "A2"


def _safe_sheet_name(name: str, used: set[str]) -> str:
    s = re.sub(r"[\[\]\\\/\?\*\:]", "_", str(name)).strip()[:31] or "Sheet"
    base = s
    i = 1
    while s in used:
        suffix = f"_{i}"
        s = (base[: 31 - len(suffix)]) + suffix
        i += 1
    used.add(s)
    return s


def _safe_filename(name: str) -> str:
    """Filesystem-safe filename component; preserves Azerbaijani/Turkish letters."""
    s = re.sub(r'[\\/:*?"<>|\r\n\t]+', "_", str(name)).strip().strip(".")
    return s[:120] or "sigorta"


def _get_upload(upload_id: str) -> dict:
    with STORE_LOCK:
        rec = STORE.get(upload_id)
    if not rec:
        abort(404, description="Yükləmə tapılmadı və ya vaxtı keçib. Zəhmət olmasa yenidən yükləyin.")
    return rec


# ----------------------- Routes -----------------------


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/upload", methods=["POST"])
def upload():
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error": "Fayl təqdim edilmədi"}), 400
    fname = secure_filename(f.filename) or "upload.xlsx"
    if not fname.lower().endswith(".xlsx"):
        return jsonify({"error": "Zəhmət olmasa .xlsx faylı yükləyin"}), 400

    try:
        groups, col_types, hidden_cols = _read_and_group(f)
    except Exception as e:
        return jsonify({"error": f"Fayl oxuna bilmədi: {e}"}), 400

    upload_id = uuid.uuid4().hex
    with STORE_LOCK:
        STORE[upload_id] = {
            "filename": fname,
            "groups": groups,
            "col_types": col_types,
            "hidden_cols": hidden_cols,
            "created": datetime.utcnow(),
        }

    if request.headers.get("X-Requested-With") == "fetch":
        return jsonify({"upload_id": upload_id, "redirect": url_for("results", upload_id=upload_id)})
    return redirect(url_for("results", upload_id=upload_id))


@app.route("/results/<upload_id>")
def results(upload_id):
    rec = _get_upload(upload_id)
    groups = rec["groups"]
    items = [
        {"name": name, "rows": len(df)}
        for name, df in sorted(groups.items(), key=lambda kv: (-len(kv[1]), kv[0]))
    ]
    total_rows = sum(len(df) for df in groups.values())
    return render_template(
        "results.html",
        upload_id=upload_id,
        filename=rec["filename"],
        items=items,
        total_rows=total_rows,
        total_groups=len(items),
    )


@app.route("/preview/<upload_id>/<path:name>")
def preview(upload_id, name):
    rec = _get_upload(upload_id)
    groups = rec["groups"]
    col_types = rec["col_types"]
    if name not in groups:
        abort(404, description=f"'{name}' sığortası tapılmadı.")
    df = groups[name]
    all_headers = list(df.columns)
    hidden_cols = rec.get("hidden_cols") or set()
    visible_idx = [i for i, h in enumerate(all_headers) if h not in hidden_cols]
    visible_headers = [all_headers[i] for i in visible_idx]

    max_preview = 2000
    total = len(df)
    shown = min(total, max_preview)
    rows = []
    for tup in df.head(shown).itertuples(index=False, name=None):
        rows.append([_fmt_cell(col_types.get(all_headers[i], "text"), tup[i]) for i in visible_idx])

    return render_template(
        "preview.html",
        upload_id=upload_id,
        name=name,
        headers=visible_headers,
        rows=rows,
        total=total,
        shown=shown,
        truncated=total > shown,
    )


@app.route("/download/<upload_id>/<path:name>")
def download_one(upload_id, name):
    rec = _get_upload(upload_id)
    groups = rec["groups"]
    if name not in groups:
        abort(404)
    data = _df_to_xlsx_bytes(groups[name], rec["col_types"], rec.get("hidden_cols"), sheet_name=name)
    fname = f"{_safe_filename(name)}.xlsx"
    return send_file(
        io.BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


@app.route("/download_all/<upload_id>")
def download_all(upload_id):
    """ZIP archive: one separate .xlsx file per insurance."""
    rec = _get_upload(upload_id)
    groups = rec["groups"]
    col_types = rec["col_types"]
    hidden_cols = rec.get("hidden_cols")

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        used_names: set[str] = set()
        for name, df in groups.items():
            base = _safe_filename(name)
            arc = f"{base}.xlsx"
            k = 1
            while arc in used_names:
                arc = f"{base}_{k}.xlsx"
                k += 1
            used_names.add(arc)
            zf.writestr(arc, _df_to_xlsx_bytes(df, col_types, hidden_cols, sheet_name=name))
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/zip",
        as_attachment=True,
        download_name="sigorta_hesabatlari.zip",
    )


@app.route("/match", methods=["POST"])
def match():
    """Box 2: take an already-split insurance .xlsx + a price list, return a
    new .xlsx (direct download) with matched columns appended.
    """
    ins = request.files.get("insurance_file")
    pl = request.files.get("pricelist")
    if not ins or not ins.filename:
        return jsonify({"error": "Sığorta faylı təqdim edilmədi"}), 400
    if not pl or not pl.filename:
        return jsonify({"error": "Qiymət cədvəli təqdim edilmədi"}), 400
    if not ins.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Sığorta faylı .xlsx formatında olmalıdır"}), 400
    if not pl.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Qiymət cədvəli .xlsx formatında olmalıdır"}), 400

    try:
        data = _match_existing_file(ins, pl)
    except Exception as e:
        return jsonify({"error": f"İşləmə xətası: {e}"}), 400

    base = _safe_filename(secure_filename(ins.filename).rsplit(".", 1)[0] or "sigorta")
    fname = f"{base}_qiymetli.xlsx"
    return send_file(
        io.BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


@app.route("/clear/<upload_id>", methods=["POST"])
def clear(upload_id):
    with STORE_LOCK:
        STORE.pop(upload_id, None)
    return redirect(url_for("index"))


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    host = "0.0.0.0" if os.environ.get("PORT") else "127.0.0.1"
    app.run(host=host, port=port, debug=debug)
